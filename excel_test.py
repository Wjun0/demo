import random

import chardet

def generate_csv():
    import csv
    first_name = ['张', '王', '李', '赵', '钱', '孙', '周', '吴']
    list_name = ['消', '美', '通过', '国', '防', '授', '权', '法', '案', '涉', '通', '过', '向', '终', '端', '用', '户', '提', '供',
                 '一', '个', '操', '作', '系', '统', '的', '多', '个', '实', '例']
    csv_file = open('tttt.csv', 'w', newline='',encoding='utf-8')
    write = csv.writer(csv_file)
    write.writerow(['id', 'name','first_name', 'last_name'])
    for i in range(1060000):
        write.writerow([i,random.choice(first_name) + random.choice(list_name), random.choice(first_name),random.choice(list_name)])
    csv_file.close()

def openpyxl_generate_excel():
    from openpyxl import Workbook
    first_name = ['张', '王', '李', '赵', '钱', '孙', '周', '吴']
    list_name = ['消', '美', '通过', '国', '防', '授', '权', '法', '案', '涉', '通', '过', '向', '终', '端', '用', '户', '提', '供',
                 '一', '个', '操', '作', '系', '统', '的', '多', '个', '实', '例']
    wb = Workbook()
    ws = wb['Sheet']
    ws.title = '王军测试'
    header = ['id', 'name','first_name', 'last_name']
    for i in range(len(header)):
        c = i + 1
        ws.cell(row=1, column=c).value = header[i]
    for i in range(1060000):
        ws.append([i,random.choice(first_name) + random.choice(list_name), random.choice(first_name),random.choice(list_name)])
    wb.save('wj_test.xlsx')
    print('success')


def generate_excel():
    from xlwt import Worksheet, Workbook
    first_name = ['张','王','李','赵','钱','孙','周','吴']
    list_name = ['消','美','通过','国','防','授','权','法','案','涉','通','过','向','终','端','用','户','提','供',
                 '一','个','操','作','系','统','的','多','个','实','例']

    ws = Workbook(encoding='utf-8')
    w = ws.add_sheet(u'sheel1')
    w.write(0, 0, 'id')
    w.write(0, 1, 'name')
    w.write(0, 2, 'first_name')
    w.write(0, 3, 'last_name')
    excel_raw = 1
    for i in range(100000):
        w.write(excel_raw,0,i)
        w.write(excel_raw,1,random.choice(first_name) + random.choice(list_name))
        w.write(excel_raw,2,random.choice(first_name))
        w.write(excel_raw,3,random.choice(list_name))
        excel_raw += 1
    ws.save('test.xls')

def open_excel():
    encoding = 'utf-8'
    with open('test.xls', 'rb')as f:
        data = f.read()
        res = chardet.detect(data)
        encoding = res['encoding']
    with open('test.xls', 'r', encoding=encoding)as f:
        lines = f.readlines()
        for i in lines:
            print(i)

def pd_open_excel():
    pass
    import pandas as pd
    df = pd.read_excel('test.xls')
    print(df)


if __name__ == '__main__':
    pass
    # generate_excel()
    # open_excel()
    # pd_open_excel()
    # openpyxl_generate_excel()
    generate_csv()